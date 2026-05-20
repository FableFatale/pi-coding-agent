import openpyxl
import os

filepath = r'D:\pi-coding-agent\promptflow\data\真实对话\301564-94-公众话术（新系统）-已完成外呼记录_2026-04-15_09-47-38.csv'

print(f"文件大小: {os.path.getsize(filepath)} bytes")

wb = openpyxl.load_workbook(filepath, read_only=True)
sheet = wb.active
print(f"总行数: {sheet.max_row}")
print(f"总列数: {sheet.max_column}")

# 获取表头
headers = [cell.value for cell in sheet[1]]
print(f"\n列名: {headers}")

# 读取前5行数据
print("\n前5行数据:")
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    if i < 6:
        print(f"Row {i}: {row}")
    else:
        break

wb.close()
